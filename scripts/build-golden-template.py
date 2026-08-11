from pathlib import Path
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1] / "docs"

rows = [
    ["id", "question", "question_type", "expected_filenames", "split", "answer_hint", "notes", "sumber"],
    ["G001", "Berapa hari cuti tahunan yang didapat karyawan penuh waktu?", "langsung", "01-kebijakan-cuti.md", "test", "12 hari", "Contoh sample — ganti sesuai corpus skripsi", "Sample docs Lumen (lib/sample-docs.ts) + golden sample"],
    ["G002", "Berapa lama jatah cuti istirahat tahunan untuk pegawai full-time?", "sinonim", "01-kebijakan-cuti.md", "test", "12 hari", "Parafrase dari G001", "Sample docs Lumen"],
    ["G003", "Jika karyawan sakit dua hari atau lebih, apa dokumen yang wajib dilampirkan?", "multi-kalimat", "01-kebijakan-cuti.md", "dev", "surat dokter", "", "Sample docs Lumen"],
    ["G004", "Apa SLA respons awal untuk insiden P1 pada jam sibuk?", "langsung", "02-runbook-oncall.md", "test", "15 menit", "", "Sample docs Lumen"],
    ["G005", "Dalam berapa menit on-call harus merespons gangguan kritis di peak hours?", "sinonim", "02-runbook-oncall.md", "dev", "15 menit; peak 11.00-14.00", "", "Sample docs Lumen"],
    ["G006", "Sebutkan tahapan utama pipeline RAG pada produk Lumen.", "langsung", "03-panduan-produk.md", "test", "chunk embed retrieve generate", "", "Sample docs Lumen"],
    ["G007", "Bagaimana kebijakan remote work mingguan (kantor vs WFH)?", "langsung", "04-remote-work.md", "test", "3 kantor 2 remote", "", "Sample docs Lumen"],
    ["G008", "Apakah akses sistem internal dari luar kantor membutuhkan VPN?", "langsung", "04-remote-work.md", "dev", "wajib VPN", "", "Sample docs Lumen"],
    ["G009", "Berapa batas waktu klaim reimbursement setelah transaksi?", "langsung", "05-sop-expense.md", "test", "30 hari", "", "Sample docs Lumen"],
    ["G010", "Saluran apa saja yang sering dipakai phishing menurut panduan keamanan?", "langsung", "06-security-awareness.md", "test", "email GitHub cloud VPN", "", "Sample docs Lumen"],
    ["G011", "Berapa jumlah approval minimal sebelum merge ke main?", "langsung", "07-coding-standards.md", "dev", "1 approval", "", "Sample docs Lumen"],
    ["G012", "Target first response time untuk tiket support prioritas tinggi?", "langsung", "08-support-playbook.md", "test", "15 menit", "", "Sample docs Lumen"],
    ["G013", "Berapa harga paket Pro bulanan?", "langsung", "09-faq-billing.md", "test", "249.000", "", "Sample docs Lumen"],
    ["G014", "Apakah ada masa trial atau garansi uang kembali di awal berlangganan?", "sinonim", "09-faq-billing.md", "dev", "7 hari pertama", "", "Sample docs Lumen"],
    ["G015", "Di lantai berapa pantri kantor berada?", "langsung", "10-office-facilities.md", "test", "lantai 12", "", "Sample docs Lumen"],
    ["G016", "Berapa lama meeting room boleh dipesan tanpa persetujuan tambahan?", "langsung", "10-office-facilities.md", "dev", "30 menit", "", "Sample docs Lumen"],
    ["G017", "Jelaskan ringkas: berapa cuti tahunan, dan apa syarat cuti sakit 2+ hari?", "multi-kalimat", "01-kebijakan-cuti.md", "test", "12 hari; surat dokter 2+ hari", "", "Sample docs Lumen"],
    ["G018", "Bandingkan SLA P1 on-call dengan first response support prioritas tinggi.", "multi-dokumen", "02-runbook-oncall.md | 08-support-playbook.md", "test", "keduanya 15 menit", "Expected lebih dari satu file", "Sample docs Lumen (2 file)"],
    ["G019", "Apa warna resmi logo perusahaan menurut brand book internal?", "tidak-ada-jawaban", "", "test", "tidak ada di corpus", "expected_filenames kosong; harus MISS / tidak tahu", "Buatan template (negatif)"],
    ["G020", "Menurut file 01-kebijakan-cuti.md berapa hari cuti tahunan?", "sebut-nama-file", "01-kebijakan-cuti.md", "dev", "12 hari", "", "Sample docs Lumen"],
    ["G021", "Bagaimana aturan cuti?", "ambigu", "01-kebijakan-cuti.md", "test", "perlu klarifikasi jenis cuti", "Query terlalu luas", "Sample docs Lumen"],
    ["G022", "Apa kepanjangan RAG pada sistem Lumen?", "langsung", "Lumen-Dokumentasi-Sistem.pdf", "test", "Retrieval Augmented Generation", "", "PDF dokumentasi Lumen (docs/)"],
    ["G023", "Berapa bobot default hybrid retrieval di Lumen untuk vector dan BM25?", "langsung", "Lumen-Dokumentasi-Sistem.pdf", "test", "0.55 vector; 0.45 BM25", "", "PDF dokumentasi Lumen"],
    ["G024", "Apakah bobot Hybrid 0.55/0.45 hasil optimasi formal atau heuristik?", "sinonim", "Lumen-Dokumentasi-Sistem.pdf", "test", "heuristik awal", "", "PDF dokumentasi Lumen / Q&A sidang"],
    ["G025", "Bagaimana Lumen mendefinisikan HIT pada evaluasi golden set?", "langsung", "Lumen-Dokumentasi-Sistem.pdf", "dev", "expected filename di top-K", "", "PDF dokumentasi Lumen"],
    ["G026", "Metrik apa yang dipakai untuk mengevaluasi retrieval di Lumen?", "langsung", "Lumen-Dokumentasi-Sistem.pdf", "test", "document-level hit; Precision@K", "", "PDF dokumentasi Lumen"],
    ["G027", "Format dokumen apa saja yang didukung Lumen?", "langsung", "Lumen-Dokumentasi-Sistem.pdf", "test", ".txt .md PDF teks", "", "PDF dokumentasi Lumen"],
    ["G028", "Mengapa PDF hasil scan kurang cocok untuk diindeks di Lumen?", "langsung", "Lumen-Dokumentasi-Sistem.pdf", "dev", "tidak ada text layer / butuh OCR", "", "PDF dokumentasi Lumen"],
    ["G029", "Apa perbedaan utama Vector retrieval dan BM25?", "langsung", "Lumen-Dokumentasi-Sistem.pdf", "test", "semantik vs leksikal", "", "PDF dokumentasi Lumen"],
    ["G030", "Mode retrieval apa yang menjadi default pada fitur chat Lumen?", "langsung", "Lumen-Dokumentasi-Sistem.pdf", "test", "Hybrid", "", "PDF dokumentasi Lumen"],
    ["G031", "Apa fokus penelitian Lumen: chatbot atau evaluasi retrieval?", "sinonim", "Lumen-Dokumentasi-Sistem.pdf", "test", "evaluasi retrieval; chat antarmuka", "", "PDF dokumentasi Lumen / framing sidang"],
    ["G032", "Sebutkan dua cara login yang didukung Lumen.", "langsung", "Lumen-Dokumentasi-Sistem.pdf", "dev", "email/password; Google OAuth", "", "PDF dokumentasi Lumen"],
    ["G033", "Jelaskan secara singkat alur ingest hingga evaluasi di Lumen.", "multi-kalimat", "Lumen-Dokumentasi-Sistem.pdf", "test", "ingest retrieve generate cite evaluate", "", "PDF dokumentasi Lumen"],
    ["G034", "Menurut Lumen-Dokumentasi-Sistem.pdf apa saja batasan sistem?", "sebut-nama-file", "Lumen-Dokumentasi-Sistem.pdf", "test", "chunk; golden selaras; bobot heuristik; LLM eksternal", "", "PDF dokumentasi Lumen"],
    ["G035", "Siapa CEO perusahaan pembuat Lumen menurut dokumen resmi?", "tidak-ada-jawaban", "", "test", "tidak tersedia", "Harus mengakui tidak ada di konteks", "Buatan template (negatif)"],
]

for i in range(36, 61):
    rows.append([f"G{i:03d}", "", "", "", "", "", "Isi dengan corpus skripsimu", ""])


def main() -> None:
    ROOT.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Golden Set"

    header_fill = PatternFill("solid", fgColor="0B1F2A")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )
    alt = PatternFill("solid", fgColor="F6F8FA")

    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, val)
            cell.alignment = wrap
            cell.border = thin
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif r_idx % 2 == 0:
                cell.fill = alt

    widths = [8, 55, 16, 36, 8, 28, 40, 42]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(rows)}"

    ws2 = wb.create_sheet("Petunjuk & Sumber")
    guide = [
        ["Topik", "Penjelasan"],
        [
            "Kenapa CSV sebelumnya berantakan?",
            "Excel Indonesia sering memakai pemisah titik-koma (;). Pakai file .xlsx ini agar kolom sudah terpisah.",
        ],
        [
            "Dari mana pertanyaan G001–G021?",
            "Dari dokumen contoh bawaan Lumen (tombol Muat contoh di Desk): cuti, on-call, produk, remote, expense, security, coding, support, billing, facilities. Mengacu lib/sample-docs.ts dan lib/eval/golden.ts.",
        ],
        [
            "Dari mana pertanyaan G022–G034?",
            "Dari docs/Lumen-Dokumentasi-Sistem.pdf (dokumentasi sistem + Q&A sidang).",
        ],
        [
            "Dari mana G019 & G035?",
            "Dibuat khusus sebagai pertanyaan negatif (jawaban tidak ada di corpus) untuk uji perilaku tidak tahu.",
        ],
        [
            "Apakah ini golden set skripsi final?",
            "BUKAN. Ini TEMPLATE CONTOH. Untuk skripsi, ganti dengan dokumen & pertanyaan domain penelitianmu.",
        ],
        [
            "Kolom question_type",
            "langsung | sinonim | multi-kalimat | multi-dokumen | tidak-ada-jawaban | sebut-nama-file | ambigu",
        ],
        [
            "Kolom expected_filenames",
            "Harus PERSIS sama dengan nama file di pustaka Desk. Multi-file pisahkan dengan ' | '.",
        ],
        [
            "Kolom split",
            "dev = pilih bobot/K (~30%). test = laporan akhir (~70%).",
        ],
        ["Target jumlah", "Realistis: 40–60 pertanyaan terisi. Ideal: 80–100."],
        [
            "Cara pakai di Lumen",
            "1) Index dokumen sesuai expected_filenames. 2) Buat golden di halaman Eval. 3) Jalankan evaluasi per mode/K.",
        ],
    ]
    for r_idx, row in enumerate(guide, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws2.cell(r_idx, c_idx, val)
            cell.alignment = wrap
            cell.border = thin
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    ws2.column_dimensions["A"].width = 36
    ws2.column_dimensions["B"].width = 100

    out_xlsx = ROOT / "golden-set-template.xlsx"
    wb.save(out_xlsx)

    out_csv = ROOT / "golden-set-template.csv"
    with out_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerows(rows)

    print(f"WROTE {out_xlsx}")
    print(f"WROTE {out_csv}")


if __name__ == "__main__":
    main()
